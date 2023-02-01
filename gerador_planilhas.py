import openpyxl
import os
from time import sleep
import pandas as pd

# INICIANDO VARIÁVEIS
#-----------------------------------------------------------------------------
#SHEETS

sheet_1 = ''
sheet_2 = ''
sheet_3 = ''

#-----------------------------------------------------------------------------

#ESCOLHAS
continuar_programa=''
criar_nova_pagina =''
pagina_escolhida = ''
adicionar_coluna =''


#COLUNA
coluna_escolhida = ''

#COLEÇÃO DE DADOS
paginas=[]
linha_dados = []
nova_linha = []


#LAYOUT
while continuar_programa !='n':
    print('============================= BEM VINDO AO GERADOR DE PLANILHAS ===================================')

    print('Para começar vamos criar uma nova pagina dentro de uma planilha')


    sheet_1 = input('Digite o nome da pagina: ')
    paginas.append(sheet_1)

    criar_nova_pagina = input('Deseja criar nova pagina ? (s/n): ')
    
            



    # verificando escolhas
    while criar_nova_pagina != 'n':
            if criar_nova_pagina != 'n':
                    sheet_2 = input('Digite o nome da pagina: ')
                    paginas.append(sheet_2)
                    

            criar_nova_pagina = input('Deseja criar nova pagina ? (s/n): ')
            if criar_nova_pagina != 'n':
                    sheet_3 = input('Digite o nome da pagina: ')
                    paginas.append(sheet_3)  

    print(f'Suas paginas são : {list(paginas)}') 

    #atribuindo o valor que o usuario escolho como sheets do arquivo    
    workbook = openpyxl.Workbook()
    del workbook['Sheet']
    workbook.create_sheet(str(sheet_1))
    workbook.create_sheet(str(sheet_2))
    workbook.create_sheet(str(sheet_3))

    #Confirmando a pagina a ser manipulada
    print('================ ============================================================= ====================')

    pagina_escolhida = input('Digite a pagina que deseja inserir dados :')
    
    print('================ ============================================================= ====================')

    if pagina_escolhida == sheet_1:
        print(f"Voce escolheu a pagina: {sheet_1}")
    elif pagina_escolhida == sheet_2:
        print(f"Voce escolheu a pagina: {sheet_2}")
    elif pagina_escolhida == sheet_3:
        print(f"Voce escolheu a pagina: {sheet_3}")
    else:
        print('Pagina não encontrada')  

   # atribuindo a planilha escolhida como workbook       
    if   sheet_1!= None:
         sheet_sheet1 =workbook[str(sheet_1)]
    elif sheet_2 !=None:    
         sheet_sheet2 =workbook[str(sheet_2)]
    elif sheet_3 != None:    
         sheet_sheet3 =workbook[str(sheet_3)]

    # nomeando o arquivo
    print('================ ============================================================= ====================')
    arquivo_nome = input('Escolha um nome para o arquivo: ') 
    print('================ ============================================================= ====================')

    print('================ ============================================================= ====================')

    print('Insira dados iniciando pelo cabeçalho ')

    print('================ ============================================================= ====================')

    #adicionando itens a sheet escolhida
          
    add_linha ='s'

    while add_linha !='n':
        if add_linha =='s':
            linha = input ('Digite os dados separados por virgula: ')
            lista=[]            
                   
            dados_trat= linha.split(',')

            for item in dados_trat:
                lista.append(item)
                        
            if pagina_escolhida == sheet_1:
                sheet_sheet1.append(lista)
            elif pagina_escolhida == sheet_2:
                sheet_sheet2.append(lista)
            elif pagina_escolhida == sheet_3:
                sheet_sheet3.append(lista)
            
            

        add_linha = input('Deseja adicionar mais linhas (s/n): ')

    arquivo_nome = workbook.save(arquivo_nome + '.xlsx')

    print(f'Dados foram salvos gerando o arquivo {str(arquivo_nome)}.xlxs')
    print('================ ============================================================= ====================')

    continuar_programa = input("Deseja continuar a criar planilhas? (s/n): ")

    print('================ ============================================================= ====================')

    if continuar_programa=='n':
        msg = 'FINALIZANDO PROGRAMA'
        for x in msg:
            print([x])
            sleep(.1)




     









   



 






   #criar_nova_pagina =input("Deseja criar nova pagina (s/n)")

   #--Novas paginas ?

   



    




